import openpyxl


def get_row_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    row_count = sheet.max_row
    return row_count

def get_column_count(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    column_count = sheet.max_column
    return column_count

def read_data(file, sheetname, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return (sheet.cell(row_num, column_num).value)

